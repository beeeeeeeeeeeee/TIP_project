import random
import googlemaps
import csv
import pandas as pd
from geopy import distance
import geopandas as gpd
import os
import numpy as np
import os
import boto3
from botocore.exceptions import ClientError
from openpyxl import load_workbook
import xlwt

from django.conf import settings
from django.shortcuts import render, HttpResponse, redirect
from django import forms
from django.db.models import Q

from app01 import models
from django.core.validators import RegexValidator
from django.core.exceptions import ValidationError
from django.views.decorators.csrf import csrf_exempt
from app01.utilis.encrypt import md5
from app01.utilis.code import check_code
from django.utils.safestring import mark_safe
from io import BytesIO
import json
from app01.utilis.pagination import Pagination
from django.http import JsonResponse
from datetime import datetime

# Create your views here.
sheet_name = ['User', 'Address', 'Piano', 'CPA', 'Tuning']

user_fields = ['uid', 'first_name', 'last_name', 'email', 'gender', 'phone_number']
address_fields = ['aid', 'address', 'suburb', 'postcode', 'lat', 'long']
piano_fields = ['pid', 'brand', 'model', 'sub_model', 'colour', 'type']
cpa_fields = ['cid', 'uid_id', 'aid_id', 'pid_id', 'sn', 'sales', 'co_sales', 'sold_date', 'active', 'directly_sold']
tuning_fields = ['tid', 'mid', 'cid_id', 'tuning_date', 'piano_condition']

data_column = {
    'User': user_fields,
    'Address': address_fields,
    'Piano': piano_fields,
    'CPA': cpa_fields,
    'Tuning': tuning_fields,
}

access_key = settings.ACCESS_KEY
access_secret = settings.ACCESS_SECRET
bucket_name = settings.BUCKET_NAME


def data_excel_file():
    user_queryset = models.User.objects.all()
    address_queryset = models.Address.objects.all()
    piano_queryset = models.Piano.objects.all()
    cpa_queryset = models.CPA.objects.all()
    tuning_queryset = models.Tuning.objects.all()

    data_queryset = {
        'User': user_queryset,
        'Address': address_queryset,
        'Piano': piano_queryset,
        'CPA': cpa_queryset,
        'Tuning': tuning_queryset,
    }

    wb = xlwt.Workbook()
    for name in sheet_name:
        ws = wb.add_sheet(name)
        row_num = 0
        columns = data_column[name]
        for col_num in range(len(columns)):
            # head of the table
            ws.write(row_num, col_num, columns[col_num])
        if name == "User":
            for row in data_queryset[name]:
                row_num += 1
                ws.write(row_num, 0, row.uid)
                ws.write(row_num, 1, row.first_name)
                ws.write(row_num, 2, row.last_name)
                ws.write(row_num, 3, row.email)
                ws.write(row_num, 4, row.gender)
                ws.write(row_num, 5, row.phone_number)
        elif name == "Address":
            for row in data_queryset[name]:
                row_num += 1
                ws.write(row_num, 0, row.aid)
                ws.write(row_num, 1, row.address)
                ws.write(row_num, 2, row.suburb)
                ws.write(row_num, 3, row.postcode)
                ws.write(row_num, 4, row.lat)
                ws.write(row_num, 5, row.long)
        elif name == "Piano":
            for row in data_queryset[name]:
                row_num += 1
                ws.write(row_num, 0, row.pid)
                ws.write(row_num, 1, row.brand)
                ws.write(row_num, 2, row.model)
                ws.write(row_num, 3, row.sub_model)
                ws.write(row_num, 4, row.colour)
                ws.write(row_num, 5, row.type)
        elif name == "CPA":
            for row in data_queryset[name]:
                row_num += 1
                ws.write(row_num, 0, row.cid)
                ws.write(row_num, 1, row.uid_id)
                ws.write(row_num, 2, row.aid_id)
                ws.write(row_num, 3, row.pid_id)
                ws.write(row_num, 4, row.sn)
                ws.write(row_num, 5, row.sales)
                ws.write(row_num, 6, row.co_sales)
                ws.write(row_num, 7, str(row.sold_date))
                ws.write(row_num, 8, row.active)
                ws.write(row_num, 9, row.directly_sold)
        elif name == "Tuning":
            for row in data_queryset[name]:
                row_num += 1
                ws.write(row_num, 0, row.tid)
                ws.write(row_num, 1, row.mid)
                ws.write(row_num, 2, row.cid_id)
                ws.write(row_num, 3, str(row.tuning_date))
                ws.write(row_num, 4, row.piano_condition)
    media_root = settings.MEDIA_ROOT
    media_path = os.path.join(media_root, "test.xls")
    # print(media_path)
    wb.save(media_path)


def upload_s3():
    file_name = datetime.now().strftime("%Y%m%d%H%M%S") + str(random.randint(1000, 9999)) + ".xls"
    media_root = settings.MEDIA_ROOT
    excel_path = os.path.join(media_root, "test.xls")
    client_s3 = boto3.client(
        "s3",
        aws_access_key_id=access_key,
        aws_secret_access_key=access_secret

    )
    client_s3.upload_file(
        excel_path,
        bucket_name,
        file_name
    )


class AddressModelForm(forms.ModelForm):
    class Meta:
        model = models.Address
        fields = "__all__"
        # exclude = ("lat", "long")

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, field in self.fields.items():
            field.widget.attrs = {"class": "form-control", "placeholder": field.label}


def address_list(request):
    search_value = request.POST.get("q", "")

    if search_value:
        queryset = models.Address.objects.filter(
            Q(address__icontains=search_value) | Q(suburb__icontains=search_value) |
            Q(postcode__icontains=search_value)).order_by("-aid")
        # queryset = models.Address.objects.filter(
        #     Q(address__icontains=search_value) & Q(suburb__icontains=search_value)).order_by("-aid")
        form = AddressModelForm()
        content = {
            "form": form,
            "queryset": queryset,
            "search_value": search_value
        }
        return render(request, "address_list.html", content)

    queryset = models.Address.objects.all().order_by("-aid")
    page_object = Pagination(request, queryset)

    form = AddressModelForm()
    content = {
        "form": form,
        "queryset": page_object.page_queryset,
        "page_string": page_object.html(),
        "search_value": search_value,
        "nbar": "data_management",

    }
    return render(request, "address_list.html", content)


def address_add(request):
    title = "Add New Address"
    if request.method == "GET":
        form = AddressModelForm()
        return render(request, "address_add.html", {"form": form, "title": title})

    form = AddressModelForm(data=request.POST)
    if form.is_valid():
        address = request.POST.get("address")
        suburb = request.POST.get("suburb")
        postcode = request.POST.get("postcode")
        lat = request.POST.get("lat")
        long = request.POST.get("long")
        # print(address)

        exists = models.Address.objects.filter(address=address, suburb=suburb, postcode=postcode).exists()
        if exists:
            address_id = models.Address.objects.filter(address=address, suburb=suburb, postcode=postcode).first().aid
            form = AddressModelForm()
            alert = f"The address is exist, and the Address ID is {address_id}, please double check"
            return render(request, "address_add.html", {"form": form, "alert": alert})

        # print(address, suburb, postcode, lat, long)
        form.save()
        # rewrite/create the file test.xls in media folder
        data_excel_file()
        # upload the test.xls file to aws s3 bucket
        upload_s3()
        return redirect("/address/list/")
    return render(request, "address_add.html", {"form": form, "title": title})


def address_multi(request):
    # "bulk upload"
    file_object = request.FILES.get("exc")
    # print(type(file_object))
    wb = load_workbook(file_object)
    sheet = wb.worksheets[1]
    # from the second rows min_row=2
    for row in sheet.iter_rows(min_row=2):
        aid = row[0].value
        address = row[1].value
        suburb = row[2].value
        postcode = row[3].value
        latitude = row[4].value
        longitude = row[5].value
        models.Address.objects.create(aid=aid, address=address, suburb=suburb, postcode=postcode,
                                      lat=latitude, long=longitude)

    # rewrite/create the file test.xls in media folder
    data_excel_file()
    # upload the test.xls file to aws s3 bucket
    upload_s3()
    return redirect("/address/list/")


def address_edit(request, nid):
    row_object = models.Address.objects.filter(aid=nid).first()
    title = "Edit Address"

    if request.method == "GET":
        form = AddressModelForm(instance=row_object)
        return render(request, "address_add.html", {"form": form, "title": title})

    form = AddressModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        # rewrite/create the file test.xls in media folder
        data_excel_file()
        # upload the test.xls file to aws s3 bucket
        upload_s3()
        return redirect("/address/list/")
    else:
        return render(request, "address_add.html", {"form": form, "title": title})


def address_delete(request):
    uid = request.GET.get("uid")
    exist = models.Address.objects.filter(aid=uid)

    if not exist:
        return JsonResponse({"status": False, "error": "Data not exist!"})
    models.Address.objects.filter(aid=uid).delete()
    # rewrite/create the file test.xls in media folder
    data_excel_file()
    # upload the test.xls file to aws s3 bucket
    upload_s3()
    return JsonResponse({"status": True})


class UserModelForm(forms.ModelForm):
    class Meta:
        model = models.User
        fields = "__all__"
        # exclude = ("lat", "long")

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, field in self.fields.items():
            field.widget.attrs = {"class": "form-control", "placeholder": field.label}


def user_list(request):
    search_value = request.POST.get("q", "")

    if search_value:
        queryset = models.User.objects.filter(
            Q(first_name__icontains=search_value) | Q(last_name__icontains=search_value) |
            Q(email__icontains=search_value) | Q(phone_number__icontains=search_value)
        ).order_by("-uid")
        form = UserModelForm()
        content = {
            "form": form,
            "queryset": queryset,
            "search_value": search_value
        }
        return render(request, "user_list.html", content)
    queryset = models.User.objects.all().order_by("-uid")
    page_object = Pagination(request, queryset)

    form = PianoModelForm()
    content = {
        "form": form,
        "queryset": page_object.page_queryset,
        "page_string": page_object.html()

    }
    return render(request, "user_list.html", content)


def user_add(request):
    title = "Add New User"
    if request.method == "GET":
        form = UserModelForm()
        return render(request, "add.html", {"form": form, "title": title})

    form = UserModelForm(data=request.POST)
    if form.is_valid():
        form.save()
        # rewrite/create the file test.xls in media folder
        data_excel_file()
        # upload the test.xls file to aws s3 bucket
        upload_s3()
        return redirect("/user/list/")
    return render(request, "add.html", {"form": form, "title": title})


def user_multi(request):
    # "bulk upload"
    file_object = request.FILES.get("exc")
    # print(type(file_object))
    wb = load_workbook(file_object)
    sheet = wb.worksheets[4]
    # from the second rows min_row=2
    for row in sheet.iter_rows(min_row=2):
        uid = row[0].value
        first_name = row[1].value
        last_name = row[2].value
        email = row[3].value
        gender = row[4].value
        phone_number = row[5].value
        models.User.objects.create(uid=uid, first_name=first_name, last_name=last_name, email=email,
                                   gender=gender, phone_number=phone_number)
        # print(pid, brand, model, sub_model, colour, piano_type)
    # rewrite/create the file test.xls in media folder
    data_excel_file()
    # upload the test.xls file to aws s3 bucket
    upload_s3()
    return redirect("/user/list/")


def user_edit(request, nid):
    row_object = models.User.objects.filter(uid=nid).first()
    title = "Edit User"

    if request.method == "GET":
        form = UserModelForm(instance=row_object)
        return render(request, "add.html", {"form": form, "title": title})

    form = PianoModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        # rewrite/create the file test.xls in media folder
        data_excel_file()
        # upload the test.xls file to aws s3 bucket
        upload_s3()
        return redirect("/user/list/")
    else:
        return render(request, "add.html", {"form": form, "title": title})


def user_delete(request):
    uid = request.GET.get("uid")
    exist = models.User.objects.filter(uid=uid)

    if not exist:
        return JsonResponse({"status": False, "error": "Data not exist!"})
    models.User.objects.filter(uid=uid).delete()
    # rewrite/create the file test.xls in media folder
    data_excel_file()
    # upload the test.xls file to aws s3 bucket
    upload_s3()
    return JsonResponse({"status": True})


class PianoModelForm(forms.ModelForm):
    class Meta:
        model = models.Piano
        fields = "__all__"
        # exclude = ("lat", "long")

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, field in self.fields.items():
            field.widget.attrs = {"class": "form-control", "placeholder": field.label}


def piano_list(request):
    search_value = request.POST.get("q", "")

    if search_value:
        queryset = models.Piano.objects.filter(
            Q(brand__icontains=search_value) | Q(model__icontains=search_value) | Q(
                sub_model__icontains=search_value) | Q(colour__icontains=search_value)).order_by("-pid")
        # queryset = models.Address.objects.filter(
        #     Q(address__icontains=search_value) & Q(suburb__icontains=search_value)).order_by("-aid")
        form = PianoModelForm()
        content = {
            "form": form,
            "queryset": queryset,
            "search_value": search_value
        }
        return render(request, "piano_list.html", content)
    queryset = models.Piano.objects.all().order_by("-pid")
    page_object = Pagination(request, queryset)

    form = PianoModelForm()
    content = {
        "form": form,
        "queryset": page_object.page_queryset,
        "page_string": page_object.html()

    }
    return render(request, "piano_list.html", content)


def piano_multi(request):
    # "bulk upload"
    file_object = request.FILES.get("exc")
    # print(type(file_object))
    wb = load_workbook(file_object)
    sheet = wb.worksheets[2]
    # from the second rows min_row=2
    for row in sheet.iter_rows(min_row=2):
        pid = row[0].value
        brand = row[1].value
        model = row[2].value
        sub_model = row[3].value
        colour = row[4].value
        piano_type = row[5].value
        models.Piano.objects.create(pid=pid, brand=brand, model=model, sub_model=sub_model,
                                    colour=colour, type=piano_type)
        # print(pid, brand, model, sub_model, colour, piano_type)
    # rewrite/create the file test.xls in media folder
    data_excel_file()
    # upload the test.xls file to aws s3 bucket
    upload_s3()
    return redirect("/piano/list/")


def piano_add(request):
    title = "Add New Piano"
    if request.method == "GET":
        form = PianoModelForm()
        return render(request, "add.html", {"form": form, "title": title})

    form = PianoModelForm(data=request.POST)
    if form.is_valid():
        form.save()
        # rewrite/create the file test.xls in media folder
        data_excel_file()
        # upload the test.xls file to aws s3 bucket
        upload_s3()
        return redirect("/piano/list/")
    return render(request, "add.html", {"form": form, "title": title})


def piano_edit(request, nid):
    row_object = models.Piano.objects.filter(pid=nid).first()
    title = "Edit Piano"

    if request.method == "GET":
        form = PianoModelForm(instance=row_object)
        return render(request, "add.html", {"form": form, "title": title})

    form = PianoModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        # rewrite/create the file test.xls in media folder
        data_excel_file()
        # upload the test.xls file to aws s3 bucket
        upload_s3()
        return redirect("/piano/list/")
    else:
        return render(request, "add.html", {"form": form, "title": title})


def piano_delete(request):
    uid = request.GET.get("uid")
    exist = models.Piano.objects.filter(pid=uid)

    if not exist:
        return JsonResponse({"status": False, "error": "Data not exist!"})
    models.Piano.objects.filter(pid=uid).delete()
    # rewrite/create the file test.xls in media folder
    data_excel_file()
    # upload the test.xls file to aws s3 bucket
    upload_s3()
    return JsonResponse({"status": True})


class CPAModelForm(forms.ModelForm):
    class Meta:
        model = models.CPA
        fields = "__all__"
        # fields = ["aid","pid]
        # exclude = ("lat", "long")
        # widgets = {"aid": forms.TextInput(),"pid":forms.PasswordInput()}

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, field in self.fields.items():
            field.widget.attrs = {"class": "form-control", "placeholder": field.label}


def cpa_list(request):
    search_value = request.POST.get("q", "")

    if search_value:
        queryset = models.CPA.objects.filter(
            Q(aid__address__icontains=search_value) | Q(aid__suburb__icontains=search_value) |
            Q(aid__postcode__icontains=search_value) | Q(pid__brand__icontains=search_value) |
            Q(pid__model__icontains=search_value) | Q(pid__sub_model__icontains=search_value) |
            Q(pid__colour__icontains=search_value) | Q(sn__icontains=search_value) |
            Q(uid__first_name__icontains=search_value) | Q(uid__last_name__icontains=search_value) |
            Q(sales__icontains=search_value) | Q(uid__phone_number__icontains=search_value)
        ).order_by("-cid")
        form = CPAModelForm()
        content = {
            "form": form,
            "queryset": queryset,
            "search_value": search_value
        }
        return render(request, "cpa_list.html", content)
    queryset = models.CPA.objects.all().order_by("-cid")
    page_object = Pagination(request, queryset)

    form = CPAModelForm()
    content = {
        "form": form,
        "queryset": page_object.page_queryset,
        "page_string": page_object.html()

    }
    return render(request, "cpa_list.html", content)


def cpa_multi(request):
    # "bulk upload"
    file_object = request.FILES.get("exc")
    # print(type(file_object))
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]
    # from the second rows min_row=2
    for row in sheet.iter_rows(min_row=2):
        cid = row[0].value
        uid = row[1].value
        aid = row[2].value
        pid = row[3].value
        sn = row[4].value
        sales = row[5].value
        co_sales = row[6].value
        sold_date = row[7].value
        active = row[8].value
        directly_sold = row[9].value
        models.CPA.objects.create(cid=cid, uid_id=uid, aid_id=aid, pid_id=pid, sn=sn, sales=sales, co_sales=co_sales,
                                  sold_date=sold_date, active=active, directly_sold=directly_sold)
        # print(cid, uid, aid, pid, sn, sales, co_sales, sold_date, active, directly_sold)
    # rewrite/create the file test.xls in media folder
    data_excel_file()
    # upload the test.xls file to aws s3 bucket
    upload_s3()
    return redirect("/cpa/list/")


def cpa_add(request):
    title = "Add New CPA Information"
    if request.method == "GET":
        form = CPAModelForm()
        return render(request, "add_cpa.html", {"form": form, "title": title})

    form = CPAModelForm(data=request.POST)
    if form.is_valid():
        form.save()
        # rewrite/create the file test.xls in media folder
        data_excel_file()
        # upload the test.xls file to aws s3 bucket
        upload_s3()
        return redirect("/cpa/list/")
    return render(request, "add_cpa.html", {"form": form, "title": title})


def cpa_edit(request, nid):
    row_object = models.CPA.objects.filter(cid=nid).first()
    title = "Edit CPA Information"

    if request.method == "GET":
        form = CPAModelForm(instance=row_object)
        return render(request, "add.html", {"form": form, "title": title})

    form = CPAModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        # rewrite/create the file test.xls in media folder
        data_excel_file()
        # upload the test.xls file to aws s3 bucket
        upload_s3()
        return redirect("/cpa/list/")
    else:
        return render(request, "add.html", {"form": form, "title": title})


def cpa_delete(request):
    uid = request.GET.get("uid")
    exist = models.CPA.objects.filter(cid=uid)

    if not exist:
        return JsonResponse({"status": False, "error": "Data not exist!"})
    models.CPA.objects.filter(cid=uid).delete()
    # rewrite/create the file test.xls in media folder
    data_excel_file()
    # upload the test.xls file to aws s3 bucket
    upload_s3()
    return JsonResponse({"status": True})


class TuningModelForm(forms.ModelForm):
    class Meta:
        model = models.Tuning
        fields = "__all__"
        # fields = ["aid","pid]
        # exclude = ("lat", "long")
        # widgets = {"aid": forms.TextInput(),"pid":forms.PasswordInput()}

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, field in self.fields.items():
            field.widget.attrs = {"class": "form-control", "placeholder": field.label}


def tuning_list(request):
    search_value = request.POST.get("q", "")

    if search_value:
        queryset = models.Tuning.objects.filter(
            Q(mid__icontains=search_value) | Q(cid__aid__address__icontains=search_value) |
            Q(cid__aid__suburb__icontains=search_value) | Q(cid__aid__postcode__icontains=search_value) |
            Q(cid__pid__brand__icontains=search_value) | Q(cid__pid__model__icontains=search_value) |
            Q(cid__uid__first_name__icontains=search_value) | Q(cid__uid__last_name__icontains=search_value) |
            Q(cid__uid__email__icontains=search_value) | Q(cid__uid__phone_number__icontains=search_value)
        ).order_by("-tid")
        # queryset = models.Address.objects.filter(
        #     Q(address__icontains=search_value) & Q(suburb__icontains=search_value)).order_by("-aid")
        form = TuningModelForm()
        content = {
            "form": form,
            "queryset": queryset,
            "search_value": search_value,
            "page_name": "Tuning List"
        }
        return render(request, "tuning_list.html", content)
    queryset = models.Tuning.objects.all().order_by("-tid")
    page_object = Pagination(request, queryset)

    form = TuningModelForm()
    content = {
        "form": form,
        "queryset": page_object.page_queryset,
        "page_string": page_object.html(),
        "page_name": "Tuning List"

    }
    return render(request, "tuning_list.html", content)


def tuning_multi(request):
    # "bulk upload"
    file_object = request.FILES.get("exc")
    # print(type(file_object))
    wb = load_workbook(file_object)
    sheet = wb.worksheets[3]
    # from the second rows min_row=2
    for row in sheet.iter_rows(min_row=2):
        tid = row[0].value
        mid = row[1].value
        cid = row[2].value
        tuning_date = row[3].value
        piano_condition = row[4].value
        models.Tuning.objects.create(tid=tid, mid=mid, cid_id=cid, tuning_date=tuning_date,
                                     piano_condition=piano_condition)

        # print(cid, uid, aid, pid, sn, sales, co_sales, sold_date, active, directly_sold)
    # rewrite/create the file test.xls in media folder
    data_excel_file()
    # upload the test.xls file to aws s3 bucket
    upload_s3()
    return redirect("/tuning/list/")


def tuning_add(request):
    title = "Add New Tuning Information"
    if request.method == "GET":
        form = TuningModelForm()
        return render(request, "add_tuning.html", {"form": form, "title": title})

    form = TuningModelForm(data=request.POST)
    if form.is_valid():
        form.save()
        # rewrite/create the file test.xls in media folder
        data_excel_file()
        # upload the test.xls file to aws s3 bucket
        upload_s3()
        return redirect("/tuning/list/")
    return render(request, "add_tuning.html", {"form": form, "title": title})


def tuning_edit(request, nid):
    row_object = models.Tuning.objects.filter(tid=nid).first()
    title = "Edit Tuning Information"

    if request.method == "GET":
        form = TuningModelForm(instance=row_object)
        return render(request, "add.html", {"form": form, "title": title})

    form = TuningModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        # rewrite/create the file test.xls in media folder
        data_excel_file()
        # upload the test.xls file to aws s3 bucket
        upload_s3()
        return redirect("/tuning/list/")
    else:
        return render(request, "add.html", {"form": form, "title": title})


def tuning_delete(request):
    uid = request.GET.get("uid")
    exist = models.Tuning.objects.filter(tid=uid)

    if not exist:
        return JsonResponse({"status": False, "error": "Data not exist!"})
    models.Tuning.objects.filter(tid=uid).delete()
    # rewrite/create the file test.xls in media folder
    data_excel_file()
    # upload the test.xls file to aws s3 bucket
    upload_s3()
    return JsonResponse({"status": True})


download_queryset = models.Tuning.objects.all().order_by("-tuning_date", "cid__aid__postcode")[:10]


class TuneSearchForm(forms.Form):
    max_purchase_date = forms.DateField(
        label="Less Than or Equal Purchase Date",
        widget=forms.TextInput(attrs={"class": "form-control datepicker"}),
        initial=""
    )
    min_purchase_date = forms.DateField(
        label="Greate Than or Equal Purchase Date",
        widget=forms.TextInput(attrs={"class": "form-control datepicker"}),
        initial=""

    )
    max_tuning_date = forms.DateField(
        label="Less Than or Equal Last Tuning/Service Date",
        widget=forms.TextInput(attrs={"class": "form-control datepicker"}),
        initial=""

    )
    min_tuning_date = forms.DateField(
        label="Greater Than or Equal Last Tuning/Service Date",
        widget=forms.TextInput(attrs={"class": "form-control datepicker"}),
        initial=""

    )
    suburb = forms.CharField(
        label="Suburb",
        widget=forms.TextInput(attrs={"class": "form-control"}),
        initial=""

    )
    postcode = forms.CharField(
        label="Postcode",
        widget=forms.TextInput(attrs={"class": "form-control"}),
        initial=""

    )


def map_queryset_context(queryset, form, title):
    """

    :param queryset: queryset from the tuning database
    :param form: the form in the map_test html, TuneSearchForm
    :return: context, for the variable dictionary send to map_test.html
    """
    key = settings.GOOGLE_API_KEY
    lat_list = []
    long_list = []
    # print(queryset)

    if not queryset:
        nw_latitude = -37.8275155
        nw_longitude = 144.9074201
        se_latitude = -37.938879
        se_longitude = 145.192031

    else:
        for row in queryset:
            lat_list.append(row.cid.aid.lat)
            long_list.append(row.cid.aid.long)
        lat_array = np.array(lat_list).astype(float)
        long_array = np.array(long_list).astype(float)

        lat_list = lat_array.tolist()
        long_list = long_array.tolist()
        nw_latitude = np.min(lat_array)
        nw_longitude = np.min(long_array)
        se_latitude = np.max(lat_array)
        se_longitude = np.max(long_array)
    context = {
        # 'result': result,
        'nw_latitude': nw_latitude,
        'nw_longitude': nw_longitude,
        'se_latitude': se_latitude,
        'se_longitude': se_longitude,
        'key': key,
        'lat_list': lat_list,
        'long_list': long_list,
        'title': title,
        'queryset': queryset,
        'form': form

    }
    return context


def map_test(request):
    key = settings.GOOGLE_API_KEY
    if request.method == "POST":
        form = TuneSearchForm(data=request.POST)
        if form.data['max_purchase_date'] == "":
            max_purchase_date = "2099-12-31"
        else:
            max_purchase_date = form.data['max_purchase_date']
        if form.data['min_purchase_date'] == "":
            min_purchase_date = "1949-01-01"
        else:
            min_purchase_date = form.data['min_purchase_date']
        if form.data['max_tuning_date'] == "":
            max_tuning_date = "2099-12-31"
        else:
            max_tuning_date = form.data['max_tuning_date']
        if form.data['min_tuning_date'] == "":
            min_tuning_date = "1949-01-01"
        else:
            min_tuning_date = form.data['min_tuning_date']
        suburb = form.data['suburb']
        postcode = form.data['postcode']
        result_query = (Q(cid__sold_date__lte=max_purchase_date) & Q(cid__sold_date__gte=min_purchase_date) &
                        Q(tuning_date__lte=max_tuning_date) & Q(tuning_date__gte=min_tuning_date) &
                        Q(cid__aid__suburb__icontains=suburb) & Q(cid__aid__postcode__icontains=postcode)
                        )
        queryset = models.Tuning.objects.filter(result_query).order_by("-tuning_date", "cid__aid__postcode")
        context = map_queryset_context(queryset, form, title="Filtered data")
        global download_queryset
        download_queryset = queryset
        return render(request, "map_test.html", context)
    form = TuneSearchForm()
    # print(request.method)

    queryset = models.Tuning.objects.all().order_by("-tuning_date", "cid__aid__postcode")[:10]
    context = map_queryset_context(queryset, form, title="The 10 latest pianos needed tuning")
    download_queryset = queryset
    return render(request, "map_test.html", context)


def map_download(request):
    global download_queryset

    output = []
    response = HttpResponse(content_type='text/csv')
    writer = csv.writer(response)
    # Header
    writer.writerow(['TID', 'MID', 'CID', 'AID', 'Address',
                     'Suburb', 'Postcode', 'PID', 'Piano Brand',
                     'Piano Model', 'SN Number', 'Last Tuning Date',
                     'Piano Condition'])
    for row in download_queryset:
        output.append([row.tid, row.mid, row.cid.cid, row.cid.aid.aid, row.cid.aid.address,
                       row.cid.aid.suburb, row.cid.aid.postcode, row.cid.pid.pid, row.cid.pid.brand,
                       row.cid.pid.model, row.cid.sn, row.tuning_date,
                       row.piano_condition])
    # CSV Data
    writer.writerows(output)
    return response


class TuneCheckForm(forms.Form):
    phone_number = forms.CharField(
        label="Phone Number",
        widget=forms.TextInput(attrs={"class": "form-control"}),
        initial=""
    )
    email = forms.EmailField(
        label="Email",
        widget=forms.TextInput(attrs={"class": "form-control"}),
        initial=""
    )
    name = forms.CharField(
        label="Name",
        widget=forms.TextInput(attrs={"class": "form-control"}),
        initial=""
    )
    address = forms.CharField(
        label="Address",
        widget=forms.TextInput(attrs={"class": "form-control"}),
        initial=""
    )


def tuning_check(request):
    key = settings.GOOGLE_API_KEY
    title = "Piano Tuning Booking"

    form = TuneCheckForm()
    nw_latitude = -37.8275155
    nw_longitude = 144.9074201
    se_latitude = -37.938879
    se_longitude = 145.192031
    lat_list = []
    long_list = []
    info_list = []
    context = {
        'nw_latitude': nw_latitude,
        'nw_longitude': nw_longitude,
        'se_latitude': se_latitude,
        'se_longitude': se_longitude,
        'key': key,
        'lat_list': lat_list,
        'long_list': long_list,
        'title': title,
        'queryset_cpa': [],
        'queryset_tuning': [],
        'form': form,
        'info_list': info_list

    }

    if request.method == "POST":
        form = TuneSearchForm(data=request.POST)
        phone_number = form.data['phone_number']
        email = form.data['email']
        name = form.data['name']
        address = form.data['address']

        queryset_phone = models.CPA.objects.filter(uid__phone_number__icontains=phone_number)
        queryset_email = models.CPA.objects.filter(uid__email__icontains=email)
        queryset_firstname = models.CPA.objects.filter(uid__first_name__icontains=name)
        queryset_lastname = models.CPA.objects.filter(uid__first_name__icontains=name)
        queryset_address = models.CPA.objects.filter(aid__address__icontains=address)

        queryset_name = queryset_firstname | queryset_lastname
        queryset_cpa = queryset_phone & queryset_email & queryset_address & queryset_name

        cid = []
        for query in queryset_cpa:
            cid.append(query.cid)
        queryset_tuning = models.Tuning.objects.filter(cid_id__in=cid).order_by("-tuning_date")
        # print(queryset_cpa)
        if queryset_cpa:
            for row in queryset_cpa:
                lat_list.append(row.aid.lat)
                long_list.append(row.aid.long)
                # print(row)

                str_info = "Name: %s %s,\nAddress: %s, %s, %s" % (row.uid.first_name,
                                                                  row.uid.last_name,
                                                                  row.aid.address,
                                                                  row.aid.suburb,
                                                                  row.aid.postcode,
                                                                  )
                info_list.append(str_info)

                lat_array = np.array(lat_list).astype(float)
                long_array = np.array(long_list).astype(float)
                lat_list = lat_array.tolist()
                long_list = long_array.tolist()

                nw_latitude = np.min(lat_array)
                nw_longitude = np.min(long_array)
                se_latitude = np.max(lat_array)
                se_longitude = np.max(long_array)

                context["nw_latitude"] = nw_latitude
                context["nw_longitude"] = nw_longitude
                context["se_latitude"] = se_latitude
                context["se_longitude"] = se_longitude
                context["lat_list"] = lat_list
                context["long_list"] = long_list
                context["info_list"] = info_list

        context["queryset_cpa"] = queryset_cpa
        context["queryset_tuning"] = queryset_tuning
        # print(context)

        return render(request, "tuning_check.html", context)

    return render(request, "tuning_check.html", context)


class TuneBookForm(forms.Form):
    mid = forms.CharField(
        label="MID",
        widget=forms.TextInput(attrs={"class": "form-control"}),
        initial="",
        required=False,
    )
    tuning_date = forms.DateField(
        label="Tuning Date",
        widget=forms.DateInput(attrs={"class": "form-control"}),
        initial=""
    )
    piano_condition = forms.CharField(
        label="Piano Condition",
        widget=forms.TextInput(attrs={"class": "form-control"}),
        initial="",
        required=False,
    )


def tuning_book(request, nid):
    queryset_cpa = models.CPA.objects.filter(cid=nid)
    queryset_tuning = models.Tuning.objects.filter(cid_id=nid)
    title = "New Tuning booking"
    form = TuneBookForm()
    context = {
        'queryset_cpa': queryset_cpa,
        'queryset_tuning': queryset_tuning,
        'form': form
    }
    if request.method == "GET":
        return render(request, "book.html", context)

    form = TuneBookForm(data=request.POST)
    # print(form)
    if form.is_valid():
        mid = request.POST.get("mid", "")
        cid_object = models.CPA.objects.get(cid=nid)
        tuning_date = request.POST.get("tuning_date", "")
        piano_condition = request.POST.get("piano_condition", "")
        new_tuning = models.Tuning(mid=mid, cid=cid_object,
                                   tuning_date=tuning_date,
                                   piano_condition=piano_condition)
        new_tuning.save()
        # rewrite/create the file test.xls in media folder
        data_excel_file()
        # upload the test.xls file to aws s3 bucket
        upload_s3()
        return redirect("/tuning/%s/book/" % nid)
    else:
        context["form"] = form
        return render(request, "book.html", context)


class CPANewForm(forms.Form):
    # User Part
    first_name = forms.CharField(
        label="First Name",
        initial=""
    )
    last_name = forms.CharField(
        label="Last Name",
        initial=""
    )
    email = forms.EmailField(
        label="Email",
        initial="",
        required=False,
    )
    gender_choices = (
        ("Male", "Male"),
        ("Female", "Female"),
    )
    gender = forms.ChoiceField(
        label="Gender",
        initial="Male",
        choices=gender_choices
    )
    phone_number = forms.CharField(
        label="Phone Number",
        initial="",
    )
    # Address Part
    address = forms.CharField(
        label="Address",
        initial="",
    )
    suburb = forms.CharField(
        label="Suburb",
        initial="",
    )
    postcode = forms.CharField(
        label="Postcode",
        initial="",
    )
    lat = forms.DecimalField(
        label="Latitude",
        max_digits=15,
        decimal_places=12
    )
    long = forms.DecimalField(
        label="Longitude",
        max_digits=15,
        decimal_places=12
    )
    # Piano Part
    brand = forms.CharField(
        label="Piano Brand",
        initial="",
    )
    model = forms.CharField(
        label="Piano Model",
        initial="",
    )
    sub_model = forms.CharField(
        label="Piano Sub Model",
        initial="",
        required=False
    )
    colour = forms.CharField(
        label="Piano Color",
        initial="",
        required=False
    )
    type_choices = (
        ("Upright", "Upright"),
        ("Grand", "Grand"),
    )
    type = forms.ChoiceField(
        label="Piano Type",
        required=False,
        choices=type_choices
    )
    #     cpa part
    sn = forms.CharField(
        label="SN Number",
        initial="",
        required=False
    )
    sales = forms.CharField(
        label="Sales",
        initial="",
        required=False
    )
    co_sales = forms.CharField(
        label="Co Sales",
        initial="",
        required=False
    )
    sold_date = forms.DateField(
        label="Sold Date",
        initial="",
        required=False
    )
    choices = (
        ("Yes", "Yes"),
        ("No", "No"),
    )
    active = forms.ChoiceField(
        label="Active Status",
        initial="Yes",
        choices=choices
    )
    directly_sold = forms.ChoiceField(
        label="Directly Sold",
        initial="Yes",
        choices=choices
    )

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, field in self.fields.items():
            field.widget.attrs = {"class": "form-control", "placeholder": field.label}


def cpa_new(request):
    # check update information, if anything create, need to update s3
    update_status = 0
    title = "Add New CPA Information"
    if request.method == "GET":
        form = CPANewForm
        content = {
            "title": title,
            "form": form,
        }

        return render(request, "cpa_new.html", content)

    form = CPANewForm(data=request.POST)

    if form.is_valid():
        # user
        first_name = request.POST.get("first_name")
        last_name = request.POST.get("last_name")
        email = request.POST.get("email")
        gender = request.POST.get("gender")
        phone_number = request.POST.get("phone_number")
        # address
        address = request.POST.get("address")
        suburb = request.POST.get("suburb")
        postcode = request.POST.get("postcode")
        lat = request.POST.get("lat")
        long = request.POST.get("long")
        # piano
        brand = request.POST.get("brand")
        model = request.POST.get("model")
        sub_model = request.POST.get("sub_model")
        colour = request.POST.get("colour")
        piano_type = request.POST.get("type")
        # CPA
        sn = request.POST.get("sn")
        sales = request.POST.get("sales")
        co_sales = request.POST.get("co_sales")
        sold_date = request.POST.get("sold_date")
        active = request.POST.get("active")
        directly_sold = request.POST.get("directly_sold")

        # User Validation
        exists = models.User.objects.filter(first_name=first_name, last_name=last_name,
                                            phone_number=phone_number).exists()
        if exists:
            uid = models.User.objects.filter(first_name=first_name, last_name=last_name,
                                             phone_number=phone_number).first().uid
        else:
            exists = models.User.objects.filter(phone_number=phone_number).exists()
            if exists:
                uid = models.User.objects.filter(phone_number=phone_number).first().uid
                first_name = models.User.objects.filter(phone_number=phone_number).first().first_name
                last_name = models.User.objects.filter(phone_number=phone_number).first().last_name
                email = models.User.objects.filter(phone_number=phone_number).first().email
                user_alert = f"The Phone Number is exist, The detail is uid: {uid}, phone_number: {phone_number}, first_name: {first_name}, last_name: {last_name}, email: {email}, please double check"
                return render(request, "cpa_new.html", {"form": form, "title": title, "user_alert": user_alert})
            else:
                models.User.objects.create(first_name=first_name, last_name=last_name, email=email, gender=gender,
                                           phone_number=phone_number)
                uid = models.User.objects.filter(phone_number=phone_number).first().uid
                update_status += 1

        # Address Validation
        exists = models.Address.objects.filter(address=address, suburb=suburb, postcode=postcode).exists()

        if exists:
            aid = models.Address.objects.filter(address=address, suburb=suburb, postcode=postcode).first().aid
        else:
            models.Address.objects.create(address=address, suburb=suburb, postcode=postcode, lat=lat, long=long)
            aid = models.Address.objects.filter(address=address, suburb=suburb, postcode=postcode).first().aid
            update_status += 1

        # Piano validation
        exists = models.Piano.objects.filter(brand=brand, model=model,
                                             sub_model=sub_model,
                                             colour=colour,
                                             type=piano_type
                                             ).exists()
        if exists:
            pid = models.Piano.objects.filter(brand=brand, model=model,
                                              sub_model=sub_model,
                                              colour=colour,
                                              type=piano_type
                                              ).first().pid
        else:
            models.Piano.objects.create(brand=brand, model=model,
                                        sub_model=sub_model,
                                        colour=colour,
                                        type=piano_type
                                        )
            pid = models.Piano.objects.filter(brand=brand, model=model,
                                              sub_model=sub_model,
                                              colour=colour,
                                              type=piano_type
                                              ).first().pid
            update_status += 1

        # CPA validate
        if aid and uid and pid:
            exists = models.CPA.objects.filter(uid=uid, aid=aid,
                                               pid=pid,
                                               sn=sn,
                                               sold_date=sold_date
                                               ).exists()
            if exists:
                cid = models.CPA.objects.filter(uid=uid, aid=aid,
                                                pid=pid,
                                                sn=sn,
                                                sold_date=sold_date
                                                ).first().cid
            else:
                models.CPA.objects.create(uid_id=uid, aid_id=aid, pid_id=pid, sn=sn,
                                          sales=sales, co_sales=co_sales,
                                          sold_date=sold_date, active=active,
                                          directly_sold=directly_sold
                                          )
                cid = models.CPA.objects.filter(uid=uid, aid=aid, pid=pid, sn=sn,
                                                sales=sales, co_sales=co_sales,
                                                sold_date=sold_date, active=active,
                                                directly_sold=directly_sold
                                                ).first().cid
                update_status += 1
        if update_status > 0:
            # rewrite/create the file test.xls in media folder
            data_excel_file()
            # upload the test.xls file to aws s3 bucket
            upload_s3()

        return redirect(f"/tuning/{cid}/book/")

    return render(request, "cpa_new.html", {"form": form, "title": title})


class SelectUserForm(forms.Form):
    phone_number = forms.CharField(
        label="Phone Number",
        initial=""
    )
    email = forms.EmailField(
        label="Email",
        initial=""
    )
    name = forms.CharField(
        label="First Name/Last Name",
        initial=""
    )

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, field in self.fields.items():
            field.widget.attrs = {"class": "form-control", "placeholder": field.label}


def select_user(request):
    form_add = UserModelForm()

    form_search = SelectUserForm()
    context = {
        'queryset_user': [],
        'form_search': form_search,
        'form_add': form_add,
        "page_name": "Select User"

    }

    if request.method == "POST":
        if request.POST["check"] == "add":
            form_add = UserModelForm(data=request.POST)
            context["form_add"] = form_add
            if form_add.is_valid():
                form_add.save()
                uid = models.User.objects.filter(phone_number=request.POST.get("phone_number")).first().uid
                # rewrite/create the file test.xls in media folder
                data_excel_file()
                # upload the test.xls file to aws s3 bucket
                upload_s3()
                return redirect(f"/select/{uid}/address/")
            else:

                return render(request, "select_user.html", context)
        else:
            form = SelectUserForm(data=request.POST)
            phone_number = form.data['phone_number']
            email = form.data['email']
            name = form.data['name']
            queryset_phone = models.User.objects.filter(phone_number__icontains=phone_number)
            queryset_email = models.User.objects.filter(email__icontains=email)
            queryset_firstname = models.User.objects.filter(first_name__icontains=name)
            queryset_lastname = models.User.objects.filter(last_name__icontains=name)
            queryset_user = queryset_phone & queryset_email & (queryset_firstname | queryset_lastname)

            # result_query = (
            #         Q(phone_number__icontains=phone_number) & Q(email__icontains=email) &
            #         Q(first_name__icontains=name) & Q(last_name__icontains=name)
            # )
            # queryset_user = models.User.objects.filter(result_query).order_by("-uid")
            context['queryset_user'] = queryset_user

            return render(request, "select_user.html", context)

    return render(request, "select_user.html", context)


def select_address(request, nid):
    uid = nid
    form_add = AddressModelForm()
    queryset = models.CPA.objects.filter(uid_id=uid)
    queryset_user = models.User.objects.filter(uid=uid)

    context = {
        'queryset': queryset,
        'queryset_user': queryset_user,
        'form_add': form_add,
        'uid': uid,
        "page_name": "Select Address"

    }
    if request.method == "POST":
        form_add = AddressModelForm(data=request.POST)
        context["form_add"] = form_add
        if form_add.is_valid():
            result_query = (
                    Q(address=request.POST.get("address")) & Q(suburb=request.POST.get("suburb")) &
                    Q(postcode=request.POST.get("postcode"))
            )
            exists = models.Address.objects.filter(result_query).exists()
            if exists:
                aid = models.Address.objects.filter(result_query).first().aid
                # print(aid)
            else:
                form_add.save()
                aid = models.Address.objects.filter(result_query).first().aid
                # rewrite/create the file test.xls in media folder
                data_excel_file()
                # upload the test.xls file to aws s3 bucket
                upload_s3()
                # print(aid)

            return redirect(f"/select/{uid}/{aid}/piano/")

    return render(request, "select_address.html", context)


class SelectPianoForm(forms.Form):
    brand = forms.CharField(
        label="Brand",
        initial=""
    )
    model = forms.CharField(
        label="Piano Model",
        initial=""
    )
    sub_model = forms.CharField(
        label="Piano Sub Model",
        initial="",
        required=False

    )
    colour = forms.CharField(
        label="Piano Color",
        initial="",
        required=False

    )
    choices = (
        ("Upright", "Upright"),
        ("Grand", "Grand"),
        ("", ""),

    )
    type = forms.ChoiceField(
        label="Piano Type",
        initial="",
        choices=choices,
        required=False
    )

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, field in self.fields.items():
            field.widget.attrs = {"class": "form-control", "placeholder": field.label}


def select_piano(request, uid, aid):
    uid = uid
    aid = aid
    # print(uid)
    # print(aid)
    form_add = PianoModelForm()

    form_search = SelectPianoForm()
    queryset_user = models.User.objects.filter(uid=uid)
    queryset_address = models.Address.objects.filter(aid=aid)
    queryset_piano = models.CPA.objects.filter(uid_id=uid)
    piano_id_list = []
    queryset_piano_list = []
    for query in queryset_piano:
        piano_id_list.append(query.pid_id)
    for i in piano_id_list:
        queryset_piano_list.append(models.Piano.objects.filter(pid=i).first())

    context = {
        'queryset_piano': queryset_piano_list,
        'queryset_user': queryset_user,
        'queryset_address': queryset_address,
        'form_search': form_search,
        'form_add': form_add,
        "uid": uid,
        'aid': aid,
        "page_name": "Select Piano"

    }

    if request.method == "POST":
        if request.POST["check"] == "add":
            form_add = PianoModelForm(data=request.POST)
            context["form_add"] = form_add
            if form_add.is_valid():
                form_add.save()
                pid = models.Piano.objects.filter(
                    brand=request.POST.get("brand"),
                    model=request.POST.get("model"),
                    sub_model=request.POST.get("sub_model", ""),
                    colour=request.POST.get("colour", ""),
                    type=request.POST.get("type", ""),
                ).first().pid
                # rewrite/create the file test.xls in media folder
                data_excel_file()
                # upload the test.xls file to aws s3 bucket
                upload_s3()
                return redirect(f"/select/{uid}/{aid}/piano/")
            else:
                return render(request, "select_piano.html", context)
        else:
            form = SelectPianoForm(data=request.POST)
            if form.data['brand'] == "":
                queryset_brand = models.Piano.objects.all()
            else:
                queryset_brand = models.Piano.objects.filter(brand__icontains=form.data['brand'])
            if form.data['model'] == "":
                queryset_model = models.Piano.objects.all()
            else:
                queryset_model = models.Piano.objects.filter(model__icontains=form.data['model'])
            if form.data['sub_model'] == "":
                queryset_sub_model = models.Piano.objects.all()
            else:
                queryset_sub_model = models.Piano.objects.filter(sub_model__icontains=form.data['sub_model'])
            if form.data['colour'] == "":
                queryset_colour = models.Piano.objects.all()
            else:
                queryset_colour = models.Piano.objects.filter(colour__icontains=form.data['colour'])
            if form.data['type'] == "":
                queryset_type = models.Piano.objects.all()
            else:
                queryset_type = models.Piano.objects.filter(type__icontains=form.data['type'])

            query = queryset_brand & queryset_model & queryset_sub_model & queryset_colour & queryset_type

            context["queryset_piano"] = query
            context["form_search"] = form

            return render(request, "select_piano.html", context)

    return render(request, "select_piano.html", context)


class AddCPAForm(forms.Form):
    sn = forms.CharField(
        label="SN",
        initial="",
        required=False

    )
    sales = forms.CharField(
        label="Sales",
        initial="",
        required=False

    )
    co_sales = forms.CharField(
        label="Co Sales",
        initial="",
        required=False

    )
    sold_date = forms.DateField(
        label="Sold Date",
        initial=datetime.now().strftime("%Y-%m-%d"),
    )
    choices = (
        ("Yes", "Yes"),
        ("No", "No"),
    )
    active = forms.ChoiceField(
        label="Active",
        initial="Yes",
        choices=choices,
    )
    directly_sold = forms.ChoiceField(
        label="Directly Sold",
        initial="Yes",
        choices=choices,
    )

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, field in self.fields.items():
            field.widget.attrs = {"class": "form-control", "placeholder": field.label}


def select_cpa(request, uid, aid, pid):
    uid = uid
    aid = aid
    pid = pid
    queryset_user = models.User.objects.filter(uid=uid)
    queryset_address = models.Address.objects.filter(aid=aid)
    queryset_piano = models.Piano.objects.filter(pid=pid)
    form_add = AddCPAForm()
    context = {
        'form_add': form_add,
        "uid": uid,
        'aid': aid,
        "pid": pid,
        "queryset_user": queryset_user,
        "queryset_address": queryset_address,
        "queryset_piano": queryset_piano,
        "page_name": "Select CPA"

    }

    if request.method == "POST":
        form_add = AddCPAForm(data=request.POST)
        context["form_add"] = form_add
        if form_add.is_valid():
            sn = form_add.data["sn"]
            sales = form_add.data["sales"]
            co_sales = form_add.data["co_sales"]
            sold_date = form_add.data["sold_date"]
            active = form_add.data["active"]
            directly_sold = form_add.data["directly_sold"]

            cid = models.CPA.objects.create(uid_id=uid, aid_id=aid, pid_id=pid,
                                            sn=sn, sales=sales, co_sales=co_sales,
                                            sold_date=sold_date, active=active,
                                            directly_sold=directly_sold).cid

            # rewrite/create the file test.xls in media folder
            data_excel_file()
            # upload the test.xls file to aws s3 bucket
            upload_s3()
            return redirect(f"/select/{cid}/check/")
        else:
            return render(request, "select_cpa.html", context)

    return render(request, "select_cpa.html", context)


def select_check(request, cid):
    cid = cid
    uid = models.CPA.objects.filter(cid=cid).first().uid_id
    aid = models.CPA.objects.filter(cid=cid).first().aid_id
    pid = models.CPA.objects.filter(cid=cid).first().pid_id

    queryset_user = models.User.objects.filter(uid=uid)
    queryset_address = models.Address.objects.filter(aid=aid)
    queryset_piano = models.Piano.objects.filter(pid=pid)
    queryset_cpa = models.CPA.objects.filter(cid=cid)

    context = {
        "queryset_user": queryset_user,
        "queryset_address": queryset_address,
        "queryset_piano": queryset_piano,
        "queryset_cpa": queryset_cpa,
        'cid': cid,
        "page_name": "Select Check"

    }

    return render(request, "select_check.html", context)


def chart_list(request):
    return render(request, "chart_list.html")


def bar_m1():
    year_query = models.CPA.objects.dates('sold_date', 'year')
    year_list = []
    year_data_list = []
    for i in year_query:
        year = i.year
        year_list.append(str(i.year))
        year_data = []
        for month in range(1,13):
            queryset = models.CPA.objects.filter(sold_date__year=year,sold_date__month=month)
            count = queryset.count()
            year_data.append(count)
        year_data_list.append(year_data)
    data_list = []
    for idx,year in enumerate(year_list):
        element = {
            'name': year,
            'type': 'bar',
            'data': year_data_list[idx]
        }
        data_list.append(element)
    legend = year_list
    month_list = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                  "Jul", "Aug","Sep","Oct","Nov","Dec"]
    result = {
        "status": True,
        "data": {
            'legend': legend,
            'month_list': month_list,
            'data_list': data_list
        }
    }
    return result

def bar_m2():
    year_query = models.CPA.objects.dates('sold_date', 'year')
    legend = []
    year_list = []
    year_data_list = []
    for i in year_query:
        year_list.append(str(i.year))
        count = models.CPA.objects.filter(sold_date__year=i.year).count()
        year_data_list.append(count)

    data_list = [
        {
            # 'name': year,
            'type': 'bar',
            'data': year_data_list
        }
    ]

    result = {
        "status": True,
        "data": {
            'legend': legend,
            'year_list': year_list,
            'data_list': data_list
        }
    }
    return result

def bar_m3():
    month_list = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                  "Jul", "Aug","Sep","Oct","Nov","Dec"]
    legend = []
    month_data_list = []
    for i in range(1,13):
        count = models.CPA.objects.filter(sold_date__month=i).count()
        month_data_list.append(count)

    data_list = [
        {
            # 'name': year,
            'type': 'bar',
            'data': month_data_list
        }
    ]

    result = {
        "status": True,
        "data": {
            'legend': legend,
            'month_list': month_list,
            'data_list': data_list
        }
    }
    return result

def bar_m4():
    queryset_distinct_suburb = models.CPA.objects.values_list('aid__suburb').distinct()
    suburb_list = []
    count_list = []
    for i in queryset_distinct_suburb:
        suburb = i[0]
        suburb_list.append(suburb)
        count = models.CPA.objects.filter(aid__suburb=suburb).count()
        count_list.append(count)

    count_array = np.array(count_list)
    suburb_array = np.array(suburb_list)

    order_index = (count_array * -1).argsort()

    top10_suburb_list = suburb_array[order_index][:10].tolist()
    top10_count_list = count_array[order_index][:10].tolist()

    x_axis = top10_suburb_list
    legend = []

    data_list = [
        {
            # 'name': year,
            'type': 'bar',
            'data': top10_count_list
        }
    ]

    result = {
        "status": True,
        "data": {
            'legend': legend,
            'x_axis': x_axis,
            'data_list': data_list
        }
    }
    return result

def bar_m5():
    queryset_distinct_suburb = models.CPA.objects.values_list('aid__postcode').distinct()
    suburb_list = []
    count_list = []
    for i in queryset_distinct_suburb:
        suburb = i[0]
        suburb_list.append(suburb)
        count = models.CPA.objects.filter(aid__postcode=suburb).count()
        count_list.append(count)

    count_array = np.array(count_list)
    suburb_array = np.array(suburb_list)

    order_index = (count_array * -1).argsort()

    top10_suburb_list = suburb_array[order_index][:10].tolist()
    top10_count_list = count_array[order_index][:10].tolist()

    x_axis = top10_suburb_list
    legend = []

    data_list = [
        {
            # 'name': year,
            'type': 'bar',
            'data': top10_count_list
        }
    ]

    result = {
        "status": True,
        "data": {
            'legend': legend,
            'x_axis': x_axis,
            'data_list': data_list
        }
    }
    return result

def bar_m6():
    queryset_distinct_value = models.CPA.objects.values_list('uid').distinct()
    value_list = []
    count_list = []
    for i in queryset_distinct_value:
        value = i[0]
        value_list.append(value)
        count = models.CPA.objects.filter(uid=value).count()
        count_list.append(count)

    count_array = np.array(count_list)
    value_array = np.array(value_list)

    order_index = (count_array * -1).argsort()

    top10_value_list = value_array[order_index][:10].tolist()
    top10_count_list = count_array[order_index][:10].tolist()

    content_list = []
    for i in top10_value_list:
        content = "User ID:" + str(i)
        content_list.append(content)


    x_axis = content_list
    legend = []

    data_list = [
        {
            # 'name': year,
            'type': 'bar',
            'data': top10_count_list
        }
    ]

    result = {
        "status": True,
        "data": {
            'legend': legend,
            'x_axis': x_axis,
            'data_list': data_list
        }
    }
    return result

def bar_m7():
    queryset_distinct_values = models.CPA.objects.values_list('pid__brand').distinct()
    value_list = []
    count_list = []
    for i in queryset_distinct_values:
        value = i[0]
        value_list.append(value)
        count = models.CPA.objects.filter(pid__brand=value).count()
        count_list.append(count)

    count_array = np.array(count_list)
    value_array = np.array(value_list)

    order_index = (count_array * -1).argsort()

    top10_value_list = value_array[order_index][:10].tolist()
    top10_count_list = count_array[order_index][:10].tolist()

    x_axis = top10_value_list
    legend = []

    data_list = [
        {
            # 'name': year,
            'type': 'bar',
            'data': top10_count_list
        }
    ]

    result = {
        "status": True,
        "data": {
            'legend': legend,
            'x_axis': x_axis,
            'data_list': data_list
        }
    }
    return result

def bar_m8():
    queryset_distinct_values = models.CPA.objects.values_list('pid').distinct()
    value_list = []
    count_list = []
    for i in queryset_distinct_values:
        value = i[0]
        value_list.append(value)
        count = models.CPA.objects.filter(pid=value).count()
        count_list.append(count)

    count_array = np.array(count_list)
    value_array = np.array(value_list)

    order_index = (count_array * -1).argsort()

    top10_value_list = value_array[order_index][:10].tolist()
    top10_count_list = count_array[order_index][:10].tolist()

    x_axis_list = []

    for idx in top10_value_list:
        brand = models.Piano.objects.filter(pid=idx).first().brand
        model = models.Piano.objects.filter(pid=idx).first().model
        x_axis_list.append(brand+'-'+model)


    x_axis = x_axis_list
    legend = []

    data_list = [
        {
            # 'name': year,
            'type': 'bar',
            'data': top10_count_list
        }
    ]

    result = {
        "status": True,
        "data": {
            'legend': legend,
            'x_axis': x_axis,
            'data_list': data_list
        }
    }
    return result

def pie_mul(brand):
    queryset_distinct_values = models.CPA.objects.filter(pid__brand=brand).values_list('pid').distinct()
    value_list = []
    count_list = []
    for i in queryset_distinct_values:
        value = i[0]
        value_list.append(value)
        count = models.CPA.objects.filter(pid=value).count()
        count_list.append(count)

    count_array = np.array(count_list)
    value_array = np.array(value_list)

    order_index = (count_array * -1).argsort()

    top10_value_list = value_array[order_index].tolist()
    top10_count_list = count_array[order_index].tolist()

    x_axis_list = []

    for idx in top10_value_list:
        model = models.Piano.objects.filter(pid=idx).first().model
        x_axis_list.append(model)


    x_axis = x_axis_list
    data_list = []
    for value,name in zip(top10_count_list,x_axis_list):
        data_list.append({'value':value,'name':name})
    # print(data_list)
    legend = []

    result = {
        "status": True,
        "data": {
            'legend': legend,
            'x_axis': x_axis,
            'data_list': data_list
        }
    }
    return result


def chart_bar(request):
    bar_m1_ele = bar_m1()
    bar_m2_ele = bar_m2()
    bar_m3_ele = bar_m3()
    bar_m4_ele = bar_m4()
    bar_m5_ele = bar_m5()
    bar_m6_ele = bar_m6()
    bar_m7_ele = bar_m7()
    bar_m8_ele = bar_m8()
    pie_m9_ele = pie_mul("Yamaha")
    pie_m10_ele = pie_mul("Ritmuller")
    pie_m11_ele = pie_mul("Pearl River")
    pie_m12_ele = pie_mul("Kayserburg")

    result = {}
    result['bar_m1'] = bar_m1_ele
    result['bar_m2'] = bar_m2_ele
    result['bar_m3'] = bar_m3_ele
    result['bar_m4'] = bar_m4_ele
    result['bar_m5'] = bar_m5_ele
    result['bar_m6'] = bar_m6_ele
    result['bar_m7'] = bar_m7_ele
    result['bar_m8'] = bar_m8_ele
    result['pie_m9'] = pie_m9_ele
    result['pie_m10'] = pie_m10_ele
    result['pie_m11'] = pie_m11_ele
    result['pie_m12'] = pie_m12_ele


    return JsonResponse(result)

# postcode_data_path = os.path.join("media", "shape_file", "vic_post.shp")
# suburb_data_path = os.path.join("media", "shape_file", "vic_suburb.shp")
# postcode_data = gpd.read_file(postcode_data_path)
# suburb_data = gpd.read_file(suburb_data_path)

# gmaps = googlemaps.Client(key=key)

# result = json.dumps(gmaps.geocode(str('Stadionstraat 5, 4815 NC Breda')))
# result2 = json.loads(result)
# latitude = result2[0]['geometry']['location']['lat']
# longitude = result2[0]['geometry']['location']['lng']
