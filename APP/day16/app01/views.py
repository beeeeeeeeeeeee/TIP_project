import random
import googlemaps
from django.conf import settings
from django.shortcuts import render, HttpResponse, redirect
from django import forms
from app01 import models
from django.core.validators import RegexValidator
from django.core.exceptions import ValidationError
from django.views.decorators.csrf import csrf_exempt
from app01.utilis.encrypt import md5
from app01.utilis.code import check_code
from django.utils.safestring import mark_safe
from io import BytesIO
import json
from app01.utilis.pagination import Pagination
from django.http import JsonResponse
from datetime import datetime


# Create your views here.

def depart_list(request):
    queryset = models.Department.objects.all()
    return render(request, "depart_list.html", {"queryset": queryset})


def depart_add(request):
    if request.method == "GET":
        return render(request, "depart_add.html")
    title = request.POST.get("title")
    models.Department.objects.create(title=title)
    return redirect("/depart/list/")


def depart_delete(request):
    nid = request.GET.get('nid')
    models.Department.objects.filter(id=nid).delete()
    return redirect("/depart/list/")


def depart_edit(request, nid):
    if request.method == "GET":
        row_object = models.Department.objects.filter(id=nid).first()
        # print(row_object.id,row_object.title)

        return render(request, "depart_edit.html", {"row_object": row_object})

    title = request.POST.get("title")
    models.Department.objects.filter(id=nid).update(title=title)

    return redirect("/depart/list/")


def depart_multi(request):
    from openpyxl import load_workbook
    # "bulk upload"
    file_object = request.FILES.get("exc")
    print(type(file_object))
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]
    # from the second rows min_row=2
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        print(text)
        exists = models.Department.objects.filter(title=text).exists()
        if not exists:
            models.Department.objects.create(title=text)

    return redirect("/depart/list/")


def user_list(request):
    queryset = models.UserInfo.objects.all()
    return render(request, "user_list.html", {"queryset": queryset})


def user_add(request):
    if request.method == "GET":
        context = {
            "gender_choice": models.UserInfo.gender_choices,
            "depart_list": models.Department.objects.all()
        }

        return render(request, "user_add.html", context)

    user = request.POST.get("user")
    pwd = request.POST.get("pwd")
    age = request.POST.get("age")
    ac = request.POST.get("ac")
    ctime = request.POST.get("ctime")
    gd = request.POST.get("gd")
    dp = request.POST.get("dp")

    models.UserInfo.objects.create(
        name=user,
        password=pwd,
        age=age,
        account=ac,
        create_time=ctime,
        gender=gd,
        depart_id=dp
    )
    return redirect("/user/list/")


class UserModelForm(forms.ModelForm):
    class Meta:
        model = models.UserInfo
        fields = ["name", "password", "age", "account", "create_time", "depart", "gender"]
        # widgets = {
        #     "name": forms.TextInput(attrs={"class":"form-control"}),
        #     "password": forms.PasswordInput(attrs={"class": "form-control"}),
        #     "age": forms.NumberInput(attrs={"class": "form-control"}),
        #     "account": forms.NumberInput(attrs={"class": "form-control"}),
        #     "create_time": forms.DateInput(attrs={"class": "form-control"}),
        #     "depart": forms.TextInput(attrs={"class": "form-control"}),
        #     "gender": forms.TextInput(attrs={"class": "form-control"}),
        # }

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, field in self.fields.items():
            print(name, field)
            field.widget.attrs = {"class": "form-control", "placeholder": field.label}


def user_model_form_add(request):
    if request.method == "GET":
        form = UserModelForm()
        return render(request, "user_add2.html", {"form": form})

    form = UserModelForm(data=request.POST)
    if form.is_valid():
        print(form.cleaned_data)
        form.save()
        return redirect("/user/list/")
    else:
        return render(request, "user_add2.html", {"form": form})

    # models.UserInfo.objects.create(name=)


def user_edit(request, nid):
    row_object = models.UserInfo.objects.filter(id=nid).first()

    if request.method == "GET":
        form = UserModelForm(instance=row_object)
        return render(request, "user_edit.html", {"form": form})

    form = UserModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return redirect("/user/list/")
    else:
        return render(request, "user_edit.html", {"form", form})


def user_delete(request, nid):
    models.UserInfo.objects.filter(id=nid).delete()
    return redirect("/user/list/")


def pretty_list(request):
    # fake number
    # for i in range(300):
    #     models.PrettyNum.objects.create(mobile="159888872", price=10, level=1,status= 1)

    data_dict = {}
    value = request.GET.get("q", "")
    if value:
        data_dict["mobile__contains"] = value

    queryset = models.PrettyNum.objects.filter(**data_dict).order_by("-id")
    page_object = Pagination(request, queryset)
    page_queryset = page_object.page_queryset
    page_string = page_object.html()

    return render(request, "pretty_list.html", {"queryset": page_queryset, "page_string": page_string})


class PrettyNum(forms.ModelForm):
    mobile = forms.CharField(
        label="Phone Num",
        validators=[RegexValidator(r"^[0-9]+$", "Please enter the number"),
                    RegexValidator(r"^159[0-9]+$", "Must start with 159")]
    )

    class Meta:
        model = models.PrettyNum
        fields = ["mobile", "price", "level", "status"]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, field in self.fields.items():
            print(name, field)
            field.widget.attrs = {"class": "form-control", "placeholder": field.label}

    def clean_mobile(self):
        txt_mobile = self.cleaned_data["mobile"]
        exists = models.PrettyNum.objects.filter(mobile=txt_mobile).exists()

        if exists:
            raise ValidationError("We had this num already.")
        return txt_mobile


def pretty_add(request):
    if request.method == "GET":
        form = PrettyNum()
        return render(request, "pretty_add.html", {"form": form})

    form = PrettyNum(data=request.POST)
    if form.is_valid():
        form.save()
        return redirect("/pretty/list/")
    return render(request, "pretty_add.html", {"form": form})


class PrettyNumEdit(forms.ModelForm):
    mobile = forms.CharField(
        label="Phone Num",
        validators=[RegexValidator(r"^[0-9]+$", "Please enter the number"),
                    RegexValidator(r"^159[0-9]+$", "Must start with 159")]
    )

    class Meta:
        model = models.PrettyNum
        fields = ["mobile", "price", "level", "status"]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, field in self.fields.items():
            print(name, field)
            # field.widget.attrs = {"class": "form-control", "placeholder": field.label}
            field.widget.attrs["class"] = "form_control"
            field.widget.attrs["placeholder"] = field.label

    def clean_mobile(self):
        # self.instance.pk
        txt_mobile = self.cleaned_data["mobile"]
        exists = models.PrettyNum.objects.exclude(id=self.instance.pk).filter(mobile=txt_mobile).exists()

        if exists:
            raise ValidationError("We had this num already.")
        return txt_mobile


def pretty_edit(request, nid):
    instance = models.PrettyNum.objects.filter(id=nid).first()
    if request.method == "GET":
        form = PrettyNumEdit(instance=instance)
        return render(request, "pretty_edit.html", {"form": form})
    form = PrettyNumEdit(data=request.POST, instance=instance)
    if form.is_valid():
        form.save()
        return redirect("/pretty/list/")
    return render(request, "pretty_edit.html", {"form": form})


def pretty_delete(request, nid):
    models.PrettyNum.objects.filter(id=nid).delete()
    return redirect("/pretty/list/")


def admin_list(request):
    # get the session info compare with cookie
    info = request.session.get("info")
    print(info)
    if not info:
        return redirect("/login/")

    queryset = models.Admin.objects.all()
    context = {
        "queryset": queryset,
    }
    return render(request, "admin_list.html", context)


class AdminModelForm(forms.ModelForm):
    confirm_password = forms.CharField(
        label="Confirm Password",
        widget=forms.PasswordInput
    )

    class Meta:
        model = models.Admin
        fields = ["username", "password", "confirm_password"]
        widgets = {
            "password": forms.PasswordInput,
        }

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, field in self.fields.items():
            # print(name, field)
            field.widget.attrs = {"class": "form-control", "placeholder": field.label}

    def clean_password(self):
        pwd = self.cleaned_data.get("password")
        return md5(pwd)

    def clean_confirm_password(self):
        confirm = self.cleaned_data.get("confirm_password")
        confirm = md5(confirm)
        pwd = self.cleaned_data.get("password")
        print(confirm)
        print(pwd)
        if confirm != pwd:
            raise ValidationError("The password you input is not same.")
        return confirm


def admin_add(request):
    if request.method == "GET":
        form = AdminModelForm()

        return render(request, "admin_add.html", {"form": form})
    form = AdminModelForm(data=request.POST)
    if form.is_valid():
        print(form.cleaned_data)

        form.save()
        return redirect("/admin/list/")

    return render(request, "admin_add.html", {"form": form})


class AdminEditModelForm(forms.ModelForm):
    class Meta:
        model = models.Admin
        fields = ["username"]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, field in self.fields.items():
            # print(name, field)
            field.widget.attrs = {"class": "form-control", "placeholder": field.label}


def admin_edit(request, nid):
    row_object = models.Admin.objects.filter(id=nid).first()
    if not row_object:
        return redirect("/admin/list/")
    if request.method == "GET":
        form = AdminEditModelForm(instance=row_object)
        return render(request, "admin_edit.html", {"form": form})

    form = AdminEditModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return redirect("/admin/list/")

    return render(request, "admin_edit.html", {"form": form})


def admin_delete(request, nid):
    models.Admin.objects.filter(id=nid).delete()
    return redirect("/admin/list/")


class AdminResetModelForm(forms.ModelForm):
    confirm_password = forms.CharField(
        label="Confirm Password",
        widget=forms.PasswordInput
    )

    class Meta:
        model = models.Admin
        fields = ["password", "confirm_password"]
        widgets = {
            "password": forms.PasswordInput,
        }

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, field in self.fields.items():
            # print(name, field)
            field.widget.attrs = {"class": "form-control", "placeholder": field.label}

    def clean_password(self):
        pwd = self.cleaned_data.get("password")
        return md5(pwd)

    def clean_confirm_password(self):
        confirm = self.cleaned_data.get("confirm_password")
        confirm = md5(confirm)
        pwd = self.cleaned_data.get("password")
        print(confirm)
        print(pwd)
        if confirm != pwd:
            raise ValidationError("The password you input is not same.")
        return confirm


def admin_reset(request, nid):
    row_object = models.Admin.objects.filter(id=nid).first()
    if not row_object:
        return redirect("/admin/list/")
    title = " - {}".format(row_object.username)

    if request.method == "GET":
        form = AdminResetModelForm()
        return render(request, "admin_reset.html", {"title": title, "form": form})

    form = AdminResetModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return redirect("/admin/list/")
    return render(request, "admin_reset.html", {"title": title, "form": form})


class LoginForm(forms.Form):
    username = forms.CharField(
        label="User Name",
        widget=forms.TextInput(attrs={"class": "form-control"})
    )
    password = forms.CharField(
        label="Password",
        widget=forms.PasswordInput(attrs={"class": "form-control"}, render_value=True)
    )
    code = forms.CharField(
        label="Code",
        widget=forms.TextInput(attrs={"class": "form-control"})
    )

    def clean_password(self):
        pwd = self.cleaned_data.get("password")
        return md5(pwd)


# same like loginForm
# class LoginModelForm(forms.ModelForm)
#     class Meta:
#         model = models.Admin
#         fields = ["username", "password"]

def login(request):
    if request.method == "GET":
        form = LoginForm()
        return render(request, "login.html", {"form": form})
    form = LoginForm(data=request.POST)
    if form.is_valid():
        print(form.cleaned_data)
        user_input_code = form.cleaned_data.pop("code")
        # expire 60s so if not get the blank str
        code = request.session.get("image_code", "")
        if code.upper() != user_input_code.upper():
            form.add_error("code", "Code is wrong or expire.")
            return render(request, "login.html", {"form": form})

        # check input code
        admin_object = models.Admin.objects.filter(
            username=form.cleaned_data["username"],
            password=form.cleaned_data["password"]
        ).first()
        if not admin_object:
            form.add_error("password", "User Name or Password is wrong")
            return render(request, "login.html", {"form": form})

        #     set cookie and session key value pair
        request.session["info"] = {"id": admin_object.id, "name": admin_object.username}
        # reset session expire time to 7 days
        request.session.set_expiry(60 * 60 * 24 * 7)
        return redirect("/admin/list/")

    return render(request, "login.html", {"form": form})


def image_code(request):
    img, code_string = check_code()
    print(code_string)
    # img write in RAM
    stream = BytesIO()
    img.save(stream, "png")
    # code_string write in session
    request.session["image_code"] = code_string
    # set the expire time
    request.session.set_expiry(60)

    return HttpResponse(stream.getvalue())


def logout(request):
    request.session.clear()
    return redirect("/login/")


class TaskModelForm(forms.ModelForm):
    class Meta:
        model = models.Task
        fields = "__all__"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, field in self.fields.items():
            # print(name, field)
            field.widget.attrs = {"class": "form-control", "placeholder": field.label}


def task_list(request):
    form = TaskModelForm()
    queryset = models.Task.objects.all().order_by("-id")

    page_object = Pagination(request, queryset)
    context = {
        "form": form,
        "queryset": page_object.page_queryset,
        "page_string": page_object.html()
    }
    return render(request, "task_list.html", context)


# exempt csrf post validation
@csrf_exempt
def task_ajax(request):
    print(request.GET)
    print(request.POST)
    data_dict = {"status": True, "data": [11, 22, 33, 44]}
    json_string = json.dumps(data_dict)
    print(json_string)
    return HttpResponse(json_string)
    # same like json.dumps
    # return JsonResponse(data_dict)


@csrf_exempt
def task_add(request):
    print(request.POST)
    # use modelform validate
    form = TaskModelForm(data=request.POST)

    if form.is_valid():
        form.save()
        data_dict = {"status": True}
        return HttpResponse(json.dumps(data_dict))

    data_dict = {"status": False, "error": form.errors}
    return HttpResponse(json.dumps(data_dict))


class OrderModelForm(forms.ModelForm):
    class Meta:
        model = models.Order
        # fields = "__all__"
        exclude = ("oid", "admin")
        # admin information has already in session

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, field in self.fields.items():
            # print(name, field)
            field.widget.attrs = {"class": "form-control", "placeholder": field.label}


def order_list(request):
    queryset = models.Order.objects.all().order_by("-id")
    page_object = Pagination(request, queryset)

    form = OrderModelForm()
    content = {
        "form": form,
        "queryset": page_object.page_queryset,
        "page_string": page_object.html()

    }
    return render(request, "order_list.html", content)


@csrf_exempt
def order_add(request):
    form = OrderModelForm(data=request.POST)

    if form.is_valid():
        # without"oid"
        # 202303142156298748

        form.instance.oid = datetime.now().strftime("%Y%m%d%H%M%S") + str(random.randint(1000, 9999))
        # get admin from session
        form.instance.admin_id = request.session["info"]["id"]
        print(form.cleaned_data)
        # save to dataset
        form.save()
        return JsonResponse({"status": True})
    return JsonResponse({"status": False, "error": form.errors})

    return render(request)


def order_delete(request):
    uid = request.GET.get("uid")
    exist = models.Order.objects.filter(id=uid)

    if not exist:
        return JsonResponse({"status": False, "error": "Data not exist!"})
    models.Order.objects.filter(id=uid).delete()
    return JsonResponse({"status": True})


def order_detail(request):
    """
    way 1
    uid = request.GET.get("uid")
    row_object = models.Order.objects.filter(id=uid).first()
    if not row_object:
        return JsonResponse({"status": False, "error": "Data is not exist"})
    result = {
        "status": True,
        "data": {
            "title": row_object.title,
            "price": row_object.price,
            "status": row_object.status,
        }
    }
    return JsonResponse(result)


    """
    # way 2
    uid = request.GET.get("uid")
    row_dict = models.Order.objects.filter(id=uid).values("title", "price", "status").first()
    if not row_dict:
        return JsonResponse({"status": False, "error": "Data is not exist"})
    result = {
        "status": True,
        "data": row_dict
    }
    return JsonResponse(result)


@csrf_exempt
def order_edit(request):
    uid = request.GET.get("uid")
    # print(uid)
    row_object = models.Order.objects.filter(id=uid).first()
    if not row_object:
        return JsonResponse({"status": False, "tips": "Data is not exist"})
    form = OrderModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return JsonResponse({"status": True})
    return JsonResponse({"status": False, "error": form.errors})


def chart_list(request):
    return render(request, "chart_list.html")


def chart_bar(request):
    """
    get the bar chart data
    :param request:
    :return:
    """
    legend1 = ["Sales", "Cost"]
    series_list = [
        {
            "name": 'sales',
            "type": 'bar',
            "data": [15, 30, 46, 30, 20, 44]
        },
        {
            "name": 'cost',
            "type": 'bar',
            'data': [5, 20, 36, 10, 10, 20]
        }
    ]
    x_axis = ["Jan", "Feb", "Mar", "Apr", "May", "Jun"]

    result = {
        "status": True,
        "data": {
            "legend": legend1,
            "series_list": series_list,
            "x_axis": x_axis

        }
    }
    return JsonResponse(result)


def chart_pie(request):
    db_data_list = [
        {"value": 1048, "name": "IT Department"},
        {"value": 735, "name": "Operation Department"},
        {"value": 580, "name": "New Media Department"},
        {"value": 843, "name": "Finance Department"},
    ]
    result = {
        "status": True,
        "data": db_data_list
    }
    return JsonResponse(result)


def chart_line(request):
    legend = ["Phone", "IPAD"]
    x_axis = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', "Aug", "Sep", "Oct", "Nov", "Dec"]
    series_list = [
        {
            "name": 'Phone',
            "type": 'line',
            "stack": 'Total',
            "data": [120, 132, 101, 134, 90, 230, 210, 232, 201, 154, 190, 330]
        },
        {
            "name": 'IPAD',
            "type": 'line',
            "stack": 'Total',
            "data": [220, 182, 191, 234, 290, 330, 310, 301, 334, 390, 330, 320]
        }
    ]
    data = {"legend": legend, "x_axis": x_axis, "series_list": series_list}
    result = {
        "status": True,
        "data": data,
    }
    return JsonResponse(result)


def upload_list(request):
    if request.method == "GET":
        return render(request, "upload_list.html")
    file_object = request.FILES.get("avatar")
    print(request.POST)
    print(request.FILES)
    print(file_object.name)
    f = open("a1.png", mode="wb")
    for chunk in file_object.chunks():
        f.write(chunk)
    f.close()
    return HttpResponse("...........")


class UpForm(forms.Form):
    name = forms.CharField(label="Name")
    age = forms.IntegerField(label="Age")
    img = forms.FileField(label="Head")
    exclude_fields = ["img"]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, field in self.fields.items():
            if name in self.exclude_fields:
                continue
            # print(name, field)
            field.widget.attrs = {"class": "form-control", "placeholder": field.label}


def upload_form(request):
    import os
    title = "Upload Form"
    if request.method == "GET":
        form = UpForm()
        return render(request, "upload_form.html", {"title": title, "form": form})
    form = UpForm(data=request.POST, files=request.FILES)
    if form.is_valid():

        from django.conf import settings
        media_root = settings.MEDIA_ROOT
        # save the img to folder and get the path
        image_object = form.cleaned_data.get("img")
        # file_path = "/app01/static/img/{}".format(image_object.name)
        # window and linux is different / \
        # db_file_path save to database, file_path use to write file into folder

        media_path = os.path.join("media", image_object.name)

        print(media_path)
        f = open(media_path, mode="wb")
        for chunk in image_object.chunks():
            f.write(chunk)
        f.close()

        models.Boss.objects.create(
            name=form.cleaned_data["name"],
            age=form.cleaned_data["age"],
            img=media_path,
        )
        print(form.cleaned_data)
        return HttpResponse("..........")
    return render(request, "upload_form.html", {"form": form, "title": title})


class UpModelForm(forms.ModelForm):
    exclude_fields = ["logo"]

    class Meta:
        model = models.City
        fields = "__all__"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, field in self.fields.items():
            if name in self.exclude_fields:
                continue
            # print(name, field)
            field.widget.attrs = {"class": "form-control", "placeholder": field.label}


def upload_model_form(request):
    title = "ModelForm Upload"
    if request.method == "GET":
        form = UpModelForm()
        return render(request, "upload_form.html", {"form": form, "title": title})
    form = UpModelForm(data=request.POST, files=request.FILES)
    if form.is_valid():
        # for file, will be save into media/city, and the path save to dataset
        form.save()
        return HttpResponse("success!")


def city_list(request):
    queryset = models.City.objects.all()
    return render(request, "city_list.html", {"queryset": queryset})


def city_add(request):
    title = "Add New City"
    if request.method == "GET":
        form = UpModelForm()
        return render(request, "upload_form.html", {"form": form, "title": title})
    form = UpModelForm(data=request.POST, files=request.FILES)
    if form.is_valid():
        # for file, will be save into media/city, and the path save to dataset
        form.save()
        return redirect("/city/list/")


def map_autofill(request):
    return render(request, "map_autofill.html")


def map_autofill2(request):
    return render(request, "map_autofill2.html")


def address_list(request):
    # return render(request, "address_list.html")
    # fake number
    # for i in range(300):
    #     models.PrettyNum.objects.create(mobile="159888872", price=10, level=1,status= 1)

    data_dict = {}
    value = request.GET.get("q", "")
    if value:
        data_dict["suburb__contains"] = value

    queryset = models.Address.objects.filter(**data_dict).order_by("-aid")
    page_object = Pagination(request, queryset)
    page_queryset = page_object.page_queryset
    page_string = page_object.html()

    return render(request, "address_list.html", {"queryset": page_queryset, "page_string": page_string})


class AddressForm(forms.ModelForm):
    address1 = forms.CharField(
        label="Address for google",
        max_length=256
    )
    apartment_unit = forms.CharField(
        label="Apartment, unit, suite, or floor",
        max_length=256
    )

    city = forms.CharField(
        label="City",
        max_length=64
    )
    state_province = forms.CharField(
        label="State/Province",
        max_length=64
    )

    country = forms.CharField(
        label="Country/Region",
        max_length=64
    )

    class Meta:

        model = models.Address
        # fields = ["aid", "suburb"]
        fields = "__all__"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, field in self.fields.items():
            print(name, field)
            field.widget.attrs = {"class": "form-control", "placeholder": field.label}

    def clean_address(self):
        txt_address = self.cleaned_data["address"]
        exists = models.Address.objects.filter(address=txt_address).exists()

        if exists:
            raise ValidationError("We had this address already.")
        return txt_address


def address_add(request):
    if request.method == "GET":
        form = AddressForm()
        return render(request, "address_add.html", {"form": form})

    form = AddressForm(data=request.POST)
    if form.is_valid():
        form.save()
        return redirect("/address/list/")
    return render(request, "address_add.html", {"form": form})


class Address1ModelForm(forms.ModelForm):
    class Meta:
        model = models.Address2
        fields = "__all__"
        # exclude = ("lat", "long")

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, field in self.fields.items():
            field.widget.attrs = {"class": "form-control", "placeholder": field.label}


def address1_list(request):
    queryset = models.Address2.objects.all().order_by("-id")
    page_object = Pagination(request, queryset)

    form = Address1ModelForm()
    content = {
        "form": form,
        "queryset": page_object.page_queryset,
        "page_string": page_object.html()

    }
    return render(request, "address1_list.html", content)


def address1_add(request):
    if request.method == "GET":
        form = Address1ModelForm()
        return render(request, "address1_add.html", {"form": form})

    form = Address1ModelForm(data=request.POST)
    if form.is_valid():
        address = request.POST.get("address")
        req_unit = request.POST.get("unit")
        lat = request.POST.get("lat")
        long = request.POST.get("long")
        print(type(req_unit))
        # unit = form.unit

        exists = models.Address2.objects.filter(address=address).exists()
        if exists:
            address_id = models.Address2.objects.filter(address=address).first().id
            data_unit = models.Address2.objects.filter(address=address).first().unit
            print(type(data_unit))
            print(data_unit)
            if not data_unit:
                form = Address1ModelForm()
                alert = f"The address is exist, and the ID is {address_id}, please double check"
                return render(request, "address1_add.html", {"form": form, "alert": alert})
            else:
                if data_unit == req_unit:
                    form = Address1ModelForm()
                    alert = f"The address is exist, and the ID is {address_id}, please double check"
                    return render(request, "address1_add.html", {"form": form, "alert": alert})

        form.save()
        return redirect("/address1/list/")
    return render(request, "address1_add.html", {"form": form})


def address1_multi(request):
    from openpyxl import load_workbook
    # "bulk upload"
    file_object = request.FILES.get("exc")
    print(type(file_object))
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]
    # from the second rows min_row=2
    for row in sheet.iter_rows(min_row=2):
        address = row[2].value
        unit = row[3].value
        suburb = row[4].value
        postcode = row[5].value
        latitude = row[6].value
        longitude = row[7].value
        models.Address2.objects.create(address=address, unit=unit, suburb=suburb, postcode=postcode, lat=latitude,
                                       long=longitude)

    return redirect("/address1/list/")


def map_test1(request):
    gmaps = googlemaps.Client(key=settings.GOOGLE_API_KEY)
    key = settings.GOOGLE_API_KEY
    result = json.dumps(gmaps.geocode(str('Stadionstraat 5, 4815 NC Breda')))
    result2 = json.loads(result)
    latitude = result2[0]['geometry']['location']['lat']
    longitude = result2[0]['geometry']['location']['lng']
    context = {
        'result': result,
        'latitude': latitude,
        'longitude': longitude,
        'key': key,
        "location": [(-37.80329615499994, 145.21508975400002),
                     (-37.80523492499998, 145.2147565140001),
                     (-37.805797894999955, 145.2146599340001),
                     (-37.80643409499993, 145.21455120400003),
                     (-37.81069162499995, 145.2138242630001),
                     (-37.810706624999966, 145.21371429400006),
                     (-37.811752954999974, 145.21353613300005),
                     (-37.81175021499996, 145.2135187230001),
                     (-37.81174597499995, 145.21350146300006), ]

    }
    return render(request, "map_test1.html", context)
